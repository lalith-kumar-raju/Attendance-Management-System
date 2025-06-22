import tkinter as tk
from tkinter import messagebox
from tkcalendar import Calendar
import openpyxl
import os
from datetime import datetime
import subprocess  # To open Excel file
import smtplib
from tkinter import *
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# File path for the Excel file and background image
EXCEL_FILE = "attendance.xlsx"
BACKGROUND_IMAGE_PATH = "./Amrita.png"

# Email details
sender_email = "abcd@gmail.com"
sender_password = " " #16 digit mail passcode
subject = "Student Absence Notification"


# Attendance System Class
class AttendanceSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("College Attendance Management System")
        self.root.geometry("1100x1000")
        self.root.state('zoomed')  # Open in maximized view

        # Load and set background image
        self.bg_image = tk.PhotoImage(file=BACKGROUND_IMAGE_PATH)
        self.bg_label = tk.Label(self.root, image=self.bg_image)
        self.bg_label.place(x=0, y=0, relwidth=1, relheight=1)

        # Title label
        self.title_label = tk.Label(self.root, text="College Attendance Management", font=("Arial", 24, "bold"),
                                    bg="#4682b4", fg="white", pady=20,padx=20)
        self.title_label.pack(pady=(20, 30))

        # Frame for Student Entry (ID and Name)
        student_frame = tk.Frame(self.root, bg="#f0f8ff")
        student_frame.pack(pady=15)

        # Student ID Entry
        self.id_label = tk.Label(student_frame, text="Student ID:", font=("Arial", 14), bg="#f0f8ff")
        self.id_label.grid(row=0, column=0, padx=10, pady=10, sticky="e")
        self.id_entry = tk.Entry(student_frame, font=("Arial", 14), width=15)
        self.id_entry.grid(row=0, column=1, padx=10, pady=10)

        # Student Name Entry
        self.student_label = tk.Label(student_frame, text="Student Name:", font=("Arial", 14), bg="#f0f8ff")
        self.student_label.grid(row=0, column=2, padx=10, pady=10, sticky="e")
        self.student_entry = tk.Entry(student_frame, font=("Arial", 14), width=20)
        self.student_entry.grid(row=0, column=3, padx=10, pady=10)

        self.label_email = tk.Label(student_frame, text="Parent's Email:", font=("Arial", 14), bg="#f0f8ff")
        self.label_email.grid(row=0, column=4, padx=10, pady=10, sticky="e")
        self.entry_email = tk.Entry(student_frame, font=("Arial", 14), width=20)
        self.entry_email.grid(row=0, column=5, padx=10, pady=10)

        # Add Student Button
        self.add_student_btn = tk.Button(student_frame, text="Add Student", command=self.add_student, font=("Arial", 14),
                                         bg="#4682b4", fg="white", activebackground="#5a9bd3", width=15)
        self.add_student_btn.grid(row=0, column=6, padx=20, pady=10)

        # Frame for Date Entry and Calendar Button
        date_frame = tk.Frame(self.root, bg="#f0f8ff")
        date_frame.pack(pady=15)

        self.date_label = tk.Label(date_frame, text="Selected Date (DD-MM-YYYY):", font=("Arial", 14), bg="#f0f8ff")
        self.date_label.grid(row=0, column=0, padx=10, pady=10)
        self.date_entry = tk.Entry(date_frame, font=("Arial", 14), width=15)
        self.date_entry.grid(row=0, column=1, padx=10, pady=10)

        # Button to open Calendar
        self.date_btn = tk.Button(date_frame, text="Select Date", command=self.open_calendar, font=("Arial", 14),
                                  bg="#4682b4", fg="white", activebackground="#5a9bd3", width=15)
        self.date_btn.grid(row=0, column=2, padx=20, pady=10)

        # Listbox to display students (ID and Name)
        list_frame = tk.Frame(self.root, bg="#f0f8ff")
        list_frame.pack(pady=15)

        self.student_listbox = tk.Listbox(list_frame, font=("Times New Roman", 16), height=8, width=50, bg="#f5f5f5", selectbackground="#4682b4") #can change font style
        self.student_listbox.pack(side=tk.LEFT, padx=10)

        # Scrollbar for Listbox
        self.scrollbar = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
        self.scrollbar.config(command=self.student_listbox.yview)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.student_listbox.config(yscrollcommand=self.scrollbar.set)

        # Buttons to mark attendance
        self.present_btn = tk.Button(self.root, text="Mark Present", command=lambda: self.mark_attendance("✔️"),
                                     font=("Arial", 14), bg="#32cd32", fg="white", activebackground="#3cb371", width=20)
        self.present_btn.pack(pady=10)

        self.absent_btn = tk.Button(self.root, text="Mark Absent", command=lambda: self.mark_attendance("❌"),
                                    font=("Arial", 14), bg="#dc143c", fg="white", activebackground="#b22222", width=20)
        self.absent_btn.pack(pady=10)

        # Button to display attendance report
        self.report_btn = tk.Button(self.root, text="View Attendance Report", command=self.view_report,
                                    font=("Arial", 14), bg="#4682b4", fg="white", activebackground="#5a9bd3", width=20)
        self.report_btn.pack(pady=(15, 10))

        # Send e-mail button
        self.button_send = tk.Button(self.root, text="Send Email to parent", command=self.send_email,
                                    font=("Arial", 14), bg="#4682b4", fg="white", activebackground="#5a9bd3", width=20)
        self.button_send.pack(pady=(15, 10))

        # Button to delete student
        self.delete_student_btn = tk.Button(self.root, text="Delete Student", command=self.delete_student,
                                             font=("Arial", 14), bg="#dc143c", fg="white", activebackground="#b22222", width=20)
        self.delete_student_btn.pack(pady=(5, 10))

        # Data storage for students and their attendance
        self.students = {}  # Dictionary of student_id: [name, attendance_status]
        self.attendance_date = None
        self.load_data()

    # Load data from Excel if it exists
    def load_data(self):
        if os.path.exists(EXCEL_FILE):
            workbook = openpyxl.load_workbook(EXCEL_FILE)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                student_id = row[0]
                student_name = row[1]
                percentage = row[2] if len(row) > 2 and row[2] else "0.00%"
                parent_email = row[3] if len(row) > 3 and row[3] else ""
                self.students[student_id] = [student_name, percentage, parent_email]
            workbook.close()
            self.update_student_listbox()  # Update the Listbox


    # Save student list and attendance to Excel
    def save_data(self):
        workbook = openpyxl.Workbook() if not os.path.exists(EXCEL_FILE) else openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        if not os.path.exists(EXCEL_FILE):
            sheet.append(["Student ID", "Name", "Attendance Percentage", "Parent's Email"])  # Add headers

        for student_id, (student_name, percentage, parent_email) in self.students.items():
            if not any(sheet.cell(row=row_idx, column=1).value == student_id for row_idx in range(2, sheet.max_row + 1)):
                sheet.append([student_id, student_name, percentage, parent_email])

        workbook.save(EXCEL_FILE)
        workbook.close()


    # Function to add student
    def add_student(self):
        student_id = self.id_entry.get()
        student_name = self.student_entry.get()
        parent_email = self.entry_email.get()  # Get the parent's email

        if student_id and student_name:
            if student_id in self.students:
                messagebox.showwarning("Input Error", "Student ID already exists!")
            else:
                self.students[student_id] = [student_name, "0.00%", parent_email]  # Store parent's email
                self.update_student_listbox()
                self.id_entry.delete(0, tk.END)
                self.student_entry.delete(0, tk.END)
                self.entry_email.delete(0, tk.END)  # Clear email entry
                self.save_data()  # Save to Excel
        else:
            messagebox.showwarning("Input Error", "Please enter Student ID, Name, and Email!")

    # Update Listbox from students dictionary
    def update_student_listbox(self):
        self.student_listbox.delete(0, tk.END)
        for student_id, (student_name, percentage, email) in self.students.items():
            self.student_listbox.insert(tk.END, f"{student_id} - {student_name} ({percentage}) | {email}")

    # Open calendar for date selection
    def open_calendar(self):
        top = tk.Toplevel(self.root)
        top.title("Select Date")
        cal = Calendar(top, selectmode='day', year=datetime.now().year, month=datetime.now().month, day=datetime.now().day)
        cal.pack(pady=20)

        def grab_date():
            selected_date = cal.get_date()
            formatted_date = datetime.strptime(selected_date, "%m/%d/%y").strftime("%d-%m-%Y")
            self.date_entry.delete(0, tk.END)
            self.date_entry.insert(0, formatted_date)
            self.attendance_date = formatted_date  # Update attendance_date on each selection
            top.destroy()

        tk.Button(top, text="Select", command=grab_date, font=("Arial", 14)).pack(pady=10)


    # Function to mark attendance (Present/Absent)
    def mark_attendance(self, status):
        if not self.attendance_date:
            self.attendance_date = self.date_entry.get()
            try:
                datetime.strptime(self.attendance_date, "%d-%m-%Y")
            except ValueError:
                messagebox.showwarning("Date Error", "Please enter a valid date in DD-MM-YYYY format.")
                return

        selected_student = self.student_listbox.curselection()
        if not selected_student:
            messagebox.showwarning("Selection Error", "Please select a student from the list!")
            return

        student_id = list(self.students.keys())[selected_student[0]]
        student_name = self.students[student_id][0]

        # Open Excel and mark attendance for the student
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active

        # Find the student row
        student_row = None
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == student_id:
                student_row = row[0].row
                break

        # Insert date column if not found
        date_column = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == self.attendance_date:
                date_column = col
                break
        if not date_column:
            date_column = sheet.max_column + 1
            sheet.cell(row=1, column=date_column).value = self.attendance_date

        sheet.cell(row=student_row, column=date_column).value = status
        workbook.save(EXCEL_FILE)
        workbook.close()

        # Recalculate attendance percentage
        self.calculate_attendance_percentage()
        self.update_student_listbox()  # Update listbox to show updated percentages

        # Move selection to the next student
        next_index = (selected_student[0] + 1) % self.student_listbox.size()
        self.student_listbox.selection_clear(0, tk.END)
        self.student_listbox.selection_set(next_index)
        self.student_listbox.activate(next_index)


    # Calculate attendance percentage for each student and store in dictionary
    def calculate_attendance_percentage(self):
        for student_id in self.students:
            total_days = 0
            present_days = 0

            # Calculate attendance from Excel
            workbook = openpyxl.load_workbook(EXCEL_FILE)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2):
                if row[0].value == student_id:
                    for cell in row[2:]:  # Adjusted index based on your sheet structure
                        if cell.value == "✔️":
                            present_days += 1
                        if cell.value in ["✔️", "❌"]:
                            total_days += 1
                    break

            workbook.close()

            # Calculate percentage and update
            percentage = (present_days / total_days) * 100 if total_days > 0 else 0
            self.students[student_id][1] = f"{percentage:.2f}%"  # Store as string with two decimals
            self.update_excel_percentage(student_id, self.students[student_id][1])  # Update Excel


    # Update Excel file with attendance percentage for each student
    def update_excel_percentage(self, student_id, percentage):
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active

        # Locate student row
        student_row = None
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == student_id:
                student_row = row[0].row
                break

        # Add or update percentage column
        percentage_column = 3
        if sheet.cell(row=1, column=percentage_column).value != "Attendance Percentage":
            sheet.insert_cols(percentage_column)
            sheet.cell(row=1, column=percentage_column).value = "Attendance Percentage"
        
        sheet.cell(row=student_row, column=percentage_column).value = percentage
        workbook.save(EXCEL_FILE)
        workbook.close()

    # Delete student from the system and Excel with confirmation
    def delete_student(self):
        selected_student = self.student_listbox.curselection()
        if not selected_student:
            messagebox.showwarning("Selection Error", "Please select a student from the list!")
            return

        student_id = list(self.students.keys())[selected_student[0]]
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete student ID {student_id}?"):
            del self.students[student_id]
            self.update_student_listbox()

            # Update Excel file
            workbook = openpyxl.load_workbook(EXCEL_FILE)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2):
                if row[0].value == student_id:
                    sheet.delete_rows(row[0].row, 1)
                    break

            workbook.save(EXCEL_FILE)
            workbook.close()

    # View attendance report directly in Excel
    def view_report(self):
        subprocess.Popen(["start", EXCEL_FILE], shell=True)

    # Function to send the email
    def send_email(self):
        selected_student = self.student_listbox.curselection()
        if not selected_student:
            messagebox.showerror("Selection Error", "Please select a student from the list!")
            return
        
        student_id = list(self.students.keys())[selected_student[0]]
        parent_email = self.students[student_id][2]  # Get the parent's email from the selected student
        attendance_percentage = self.students[student_id][1]  # Get the student's attendance percentage

        if not parent_email:
            messagebox.showerror("Input Error", "No parent's email associated with this student.")
            return
        
        # Update the email message to include attendance percentage
        message_with_percentage = f"""Dear Parent,

    This is to inform you that your child, {self.students[student_id][0]}, was absent today.
    Their current attendance percentage is {attendance_percentage}.
    Please contact the school if you have any questions.

    Regards,
    School Administration
    """

        try:
            # Setup the MIME
            msg = MIMEMultipart()
            msg['From'] = sender_email
            msg['To'] = parent_email
            msg['Subject'] = subject
            msg.attach(MIMEText(message_with_percentage, 'plain'))
            
            # Create the SMTP session
            server = smtplib.SMTP('smtp.gmail.com', 587)  # Replace 'smtp.example.com' with the actual SMTP server
            server.starttls()
            server.login(sender_email, sender_password)
            text = msg.as_string()
            server.sendmail(sender_email, parent_email, text)
            server.quit()
            
            messagebox.showinfo("Success", "Email sent successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

# Main Program Execution
if __name__ == "__main__":
    root = tk.Tk()
    app = AttendanceSystem(root)
    root.mainloop()
